from openpyxl import load_workbook
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill

csv_file_path = 'Data'
# Open the CSV file in read mode
fileName = "Data.csv"
fileName2 = "Data2.csv"
dict = {}
dict2 ={}
highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

uniqueIdentifier = "CRN"
uniqueIdentifierCol = 0
with open(fileName, 'r', newline='') as csvfile:
    # Create a CSV reader object
    csv_reader = csv.reader(csvfile)




    # Iterate over each row in the CSV file
    for i, row in enumerate(csv_reader):
        if i == 0:  # This if statement finds the index of the given unique identifier in column 1
            for j, v in enumerate(row):
                if v == uniqueIdentifier:
                    uniqueIdentifierCol = j
        else:
            dict[row[uniqueIdentifierCol]] = row
# dicT[25474]= 202402,CPS,100,001,25474,,A,CLASS,D,GLSN,310,,,W,,,,,10:50 AM,12:05 PM,20,21,-1,0,0,0,1,04,Introduction to Cybersecurity,1,"Renna, James", ,  ,0,,,,
# dict[25795]= 202402,CPS,100,002,25795,,A,CLASS,D,GLSN,310,M,,,,,,,10:50 AM,12:05 PM,20,20,0,0,0,0,1,04,Introduction to Cybersecurity,1,"Renna, James", ,  ,0,,,,


#print(dict)

with open(fileName2, 'r', newline='') as csvfile2:
    # Create a CSV reader object
    csv_reader2 = csv.reader(csvfile2)
    for i, row in enumerate(csv_reader2):
        if i == 0:  # This if statement finds the index of the given unique identifier in the first column
            for j, v in enumerate(row):
                if v == uniqueIdentifier:
                    uniqueIdentifierCol = j
        else:
            dict2[row[uniqueIdentifierCol]] = row


for key in dict.keys():
    if key in dict2.keys():
        if dict[key] != dict2[key]:
            print(f"Differences found for key '{key}':")
            dict[key] += "*"
            print(f"Dictionary 1 value: {dict[key]}")
            print(f"Dictionary 2 value: {dict2[key]}")
    else:
        print(f"Key '{key}' only exists in Dictionary 1")

for key in dict2.keys():
    if key not in dict.keys():
        print(f"Key '{key}' only exists in Dictionary 2")

wb = Workbook()
ws = wb.active

# Write headers
headers = ['CRN', 'Data']
ws.append(headers)

# Write data from dict1 to the worksheet
for key, value_list in dict.items():
    # Convert each element in value_list to string and concatenate with key
    ws.append([key] + [str(val) for val in value_list])

# Save the workbook
wb.save('dictionary_data.xlsx')

wb2 = load_workbook('dictionary_data.xlsx')
ws2 = wb2.active

for row in ws2:
    for cell in row:
        if cell.value == '*':
            for x in row:
                x.fill = highlight_fill

wb2.save('final.xlsx')
fileName = "Data.csv"
fileName2 = "Data2.csv"

# Unique identifier
uniqueIdentifier = "CRN"

# Function to compare two dictionaries and highlight differences
def compare_and_highlight(dict1, dict2):
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ['CRN', 'Data']
    ws.append(headers)

    # Iterate through the dictionaries
    for key in dict1.keys():
        if key in dict2.keys():
            if dict1[key] != dict2[key]:
                print(f"Differences found for key '{key}':")
                for i, (val1, val2) in enumerate(zip(dict1[key], dict2[key])):
                    if val1 != val2:
                        dict1[key][i] = f"{val1}*"
                        dict2[key][i] = f"{val2}*"
                        ws.append([key] + [val1])
                        ws.append([key] + [val2])
                        for row in ws.iter_rows(min_row=ws.max_row-1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                            for cell in row:
                                if '*' in str(cell.value):
                                    cell.fill = highlight_fill
        else:
            print(f"Key '{key}' only exists in Dictionary 1")
            ws.append([key] + dict1[key])
            for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.fill = highlight_fill

    for key in dict2.keys():
        if key not in dict1.keys():
            print(f"Key '{key}' only exists in Dictionary 2")
            ws.append([key] + dict2[key])
            for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.fill = highlight_fill

    # Save the workbook
    wb.save('exactDifferences.xlsx')

# Function to read CSV files into dictionaries
def read_csv(file_path):
    data_dict = {}
    with open(file_path, 'r', newline='') as csvfile:
        csv_reader = csv.reader(csvfile)
        for i, row in enumerate(csv_reader):
            if i == 0:  # Find the index of the unique identifier
                uniqueIdentifierCol = row.index(uniqueIdentifier)
            else:
                data_dict[row[uniqueIdentifierCol]] = row
    return data_dict

# Read CSV files into dictionaries
data_dict1 = read_csv(fileName)
data_dict2 = read_csv(fileName2)

# Compare dictionaries and highlight differences
compare_and_highlight(data_dict1, data_dict2)








